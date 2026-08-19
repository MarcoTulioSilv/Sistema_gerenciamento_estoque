"""
Modulo_03_relatorios · xlsx_builder.py
Sprint 4 — XlsxBuilder: gera arquivos .xlsx para os 4 tipos de relatório.
Lotes vencidos destacados em vermelho (RF-22).
"""
import logging
import sys
import tempfile
from datetime import date, datetime, timedelta
from pathlib  import Path
from decimal  import Decimal
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side 
from openpyxl import Workbook
from Modulo_06_dados import get_read_session, Movimentacao, Lote, Produto, Usuario
from sqlalchemy.orm import joinedload
from datetime import datetime as dt
import io as _io_bg


logger = logging.getLogger(__name__)
  
# Paleta de cores para o XLSX
COR_HEADER_FILL  = "1F4E79"   # azul escuro
COR_HEADER_FONT  = "FFFFFF"   # branco
COR_VENCIDO_FILL = "FCEBEB"   # vermelho claro
COR_VENCIDO_FONT = "A32D2D"   # vermelho escuro
COR_LINHA_ALT    = "F2F1ED"   # cinza claro (linhas alternadas)
COR_AMBER_FILL   = "FAEEDA"   # âmbar claro
COR_AMBER_FONT   = "854F0B"   # âmbar escuro
 
# imagem de fundo
if getattr(sys, 'frozen', False):
    # Se estiver rodando como um executável empacotado (.exe)
    _BASE_DIR = Path(sys.executable).parent
else:
    # Se estiver rodando no código-fonte (.py em desenvolvimento)
    _BASE_DIR = Path(__file__).resolve().parent.parent
LOGO_BG_PATH= _BASE_DIR / "assets" / "logo_Centro_Uro_Nefrologia_sem_fundo.png"
 

MIME_MAP = {
    "png":  "image/png",
    "jpg":  "image/jpeg",
    "jpeg": "image/jpeg",
    "gif":  "image/gif",
    "tiff": "image/tiff",
}


def _aplicar_background(xlsx_bytes: bytes) -> bytes:
    """
    Insere LOGO_BG_PATH como background (marca d'água) em todas as abas.
    Retorna os bytes modificados; em caso de erro retorna xlsx_bytes intacto.
    """
    if not LOGO_BG_PATH.exists():
        logger.warning(
            "Logo de background não encontrada em '%s' — planilha sem marca d'água.",
            LOGO_BG_PATH,
        )
        return xlsx_bytes
    try:
        img_bytes = LOGO_BG_PATH.read_bytes()
        img_ext   = LOGO_BG_PATH.suffix.lstrip(".").lower()
        return _injetar_background_ooxml(xlsx_bytes, img_bytes, img_ext)
    except Exception as exc:
        logger.warning("Erro ao aplicar background: %s", exc)
        return xlsx_bytes


def _injetar_background_ooxml(
    xlsx_bytes: bytes,
    img_bytes:  bytes,
    img_ext:    str = "png",
) -> bytes:
    """
    Injeta a imagem como background OOXML em todas as abas do .xlsx.

    Três requisitos do Excel para o background funcionar:
      1. Imagem em xl/media/logo_background.<ext>
      2. <Default Extension="png" ContentType="image/png"/> no Content_Types.xml
      3. <picture r:id="rIdBG"/> dentro de <worksheet>
      4. Relationship no _rels de cada aba
    """
    import zipfile
    import io as _io

    buf_in  = _io.BytesIO(xlsx_bytes)
    buf_out = _io.BytesIO()
    REL_TYPE   = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
    MEDIA_PATH = f"xl/media/logo_background.{img_ext}"
    RID        = "rIdBG"
    MIME       = MIME_MAP.get(img_ext, "image/png")

    with zipfile.ZipFile(buf_in, "r") as zin, \
         zipfile.ZipFile(buf_out, "w", zipfile.ZIP_DEFLATED) as zout:

        nomes  = zin.namelist()
        sheets = [n for n in nomes
                  if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")]
        img_gravada = False

        for item in nomes:
            data = zin.read(item)

            # 1. Gravar imagem no media
            if not img_gravada and item == (sheets[0] if sheets else ""):
                zout.writestr(MEDIA_PATH, img_bytes)
                img_gravada = True

            # 2. Content_Types: registrar extensão PNG/JPEG
            if item == "[Content_Types].xml":
                xml = data.decode("utf-8")
                entry = f'<Default Extension="{img_ext}" ContentType="{MIME}"/>'
                if f'Extension="{img_ext}"' not in xml:
                    xml = xml.replace(
                        'xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
                        f'xmlns="http://schemas.openxmlformats.org/package/2006/content-types">' + entry,
                    )
                zout.writestr(item, xml.encode("utf-8"))

            # 3. Sheet XML: injetar <picture> e namespace r:
            elif item in sheets:
                xml = data.decode("utf-8")
                if "xmlns:r=" not in xml:
                    xml = xml.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
                    )
                if "<picture " not in xml:
                    xml = xml.replace(
                        "</worksheet>",
                        f'<picture r:id="{RID}"/></worksheet>',
                    )
                zout.writestr(item, xml.encode("utf-8"))

            # 4. _rels: adicionar relationship da imagem
            elif item.startswith("xl/worksheets/_rels/") and item.endswith(".rels"):
                xml = data.decode("utf-8")
                if RID not in xml:
                    rel = (
                        f'<Relationship Id="{RID}" Type="{REL_TYPE}" '                        f'Target="../{MEDIA_PATH[3:]}"/>'                    )
                    xml = xml.replace("</Relationships>", rel + "</Relationships>")
                zout.writestr(item, xml.encode("utf-8"))

            else:
                zout.writestr(item, data)

        # 5. Criar _rels para sheets que não tinham
        for sheet in sheets:
            rels_path = (
                sheet.replace("xl/worksheets/", "xl/worksheets/_rels/") + ".rels"
            )
            if rels_path not in nomes:
                rels_xml = (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'                    '<Relationships '                    'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'                    f'<Relationship Id="{RID}" Type="{REL_TYPE}" '                    f'Target="../{MEDIA_PATH[3:]}"/>'                    "</Relationships>"
                )
                zout.writestr(rels_path, rels_xml.encode("utf-8"))

    buf_out.seek(0)
    return buf_out.read()


def _wb_styles():
    """Retorna os estilos reutilizáveis para o workbook."""
    header_fill = PatternFill("solid", fgColor=COR_HEADER_FILL)
    header_font = Font(bold=True, color=COR_HEADER_FONT, size=10)
    header_align= Alignment(horizontal="center", vertical="center", wrap_text=True)
 
    venc_fill = PatternFill("solid", fgColor=COR_VENCIDO_FILL)
    venc_font = Font(color=COR_VENCIDO_FONT, size=10)
 
    amber_fill = PatternFill("solid", fgColor=COR_AMBER_FILL)
    amber_font = Font(color=COR_AMBER_FONT, size=10)
 
    alt_fill  = PatternFill("solid", fgColor=COR_LINHA_ALT)
    norm_font = Font(size=10)
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center")
 
    thin_border = Border(
        left   = Side(style="thin", color="D3D1C7"),
        right  = Side(style="thin", color="D3D1C7"),
        top    = Side(style="thin", color="D3D1C7"),
        bottom = Side(style="thin", color="D3D1C7"),
    )
    return {
        "hf": header_fill, "hft": header_font, "ha": header_align,
        "vf": venc_fill,   "vft": venc_font,
        "af": amber_fill,  "aft": amber_font,
        "alt": alt_fill,   "nf": norm_font,
        "center": center,  "left": left,
        "border": thin_border,
    }
 
 
def _aplicar_header(ws, colunas: list[tuple[str, int]], st: dict, row_idx: int = 1):
    """Escreve o cabeçalho na linha especificada e configura larguras de coluna."""
    for col, (titulo, largura) in enumerate(colunas, 1):
        cell = ws.cell(row=row_idx, column=col, value=titulo)
        cell.fill      = st["hf"]
        cell.font      = st["hft"]
        cell.alignment = st["ha"]
        cell.border    = st["border"]
        ws.column_dimensions[cell.column_letter].width = largura
    ws.row_dimensions[row_idx].height = 28
 
 
def _aplicar_linha(ws, row_idx: int, valores: list, st: dict,
                   fill=None, font=None, alignments=None):
    """Escreve uma linha de dados com estilo e alinhamento sob medida."""
    fundo = fill  or (st["alt"] if row_idx % 2 == 0 else None)
    fonte = font  or st["nf"]
    for col, val in enumerate(valores, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font      = fonte
        if alignments and col - 1 < len(alignments):
            cell.alignment = alignments[col - 1]
        else:
            cell.alignment = st["center"] if isinstance(val, (int, float, Decimal)) else st["left"]
            
        cell.border    = st["border"]
        if fundo:
            cell.fill = fundo
    ws.row_dimensions[row_idx].height = 18
 
 
class XlsxBuilder:
    """Gera arquivos XLSX para os 4 tipos de relatório do SCE."""
 
    # ── Diretório temporário para os arquivos gerados ──────────────────────
 
    @staticmethod
    def _dir_temp() -> Path:
        d = Path(tempfile.gettempdir())/"SCU-Uronefrologia" / "relatorios"
        d.mkdir(exist_ok=True,parents=True)
        return d
 
    @staticmethod
    def _nome_arquivo(tipo: str) -> Path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return XlsxBuilder._dir_temp() / f"SCE_uro_nefrologia_{tipo}_{ts}.xlsx"
 
    # ── 1. Movimentação por período ────────────────────────────────────────
 
    @staticmethod
    def movimentacao(data_ini: date, data_fim: date) -> Path:
        """
        Relatório de movimentações (entradas e saídas) no período com células mescladas.
        RF-15: inclui produto, lote, NF, tipo, quantidade, usuário e data.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimentação"
        st = _wb_styles()
 
        colunas = [
            ("Data/Hora", 18), ("Produto", 50), ("Lote", 14),
            ("Nota Fiscal", 14), ("Tipo", 16), ("Quantidade", 12),
            ("Usuário", 20), ("Observação", 45),
        ]
        inicio = dt.combine(data_ini, dt.min.time())
        fim    = dt.combine(data_fim, dt.max.time())
        hoje = date.today()

        with get_read_session() as s:
            movs = (
                s.query(Movimentacao)
                .join(Lote)
                .join(Produto)
                .join(Usuario)
                .options(
                    joinedload(Movimentacao.lote).joinedload(Lote.produto),
                    joinedload(Movimentacao.usuario),
                )
                .filter(Movimentacao.data_hora.between(inicio, fim))
                .order_by(Movimentacao.data_hora.desc())
                .all()
            )
            
            # --- 1. CONFIGURAÇÃO DA LINHA 1 (PERÍODO E TOTAL) ---
            ws.row_dimensions[1].height = 25
            
            # Mesclagem A1:D1 para o Período
            ws.merge_cells('A1:D1')
            cell_p = ws['A1']
            cell_p.value = f"Período: {data_ini.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
            cell_p.font = Font(bold=True, color=COR_HEADER_FILL, size=11)
            cell_p.alignment = Alignment(horizontal="center", vertical="center")

            # Mesclagem G1:H1 para o Total de Movimentações
            ws.merge_cells('G1:H1')
            cell_t = ws['G1']
            cell_t.value = f"Total: {len(movs)} registros"
            cell_t.font = Font(bold=True, color=COR_HEADER_FILL, size=11)
            cell_t.alignment = Alignment(horizontal="center", vertical="center")

            # --- 2. LINHA 2: TÍTULOS DAS COLUNAS ---
            _aplicar_header(ws, colunas, st, row_idx=2)
 
            # Definição de alinhamentos para os dados
            alinhamentos_da_linha = [
                st["center"], # Data/Hora
                st["left"],   # Produto
                st["center"], # Lote
                st["center"], # Nota Fiscal
                st["center"], # Tipo
                st["center"], # Quantidade
                st["center"], # Usuário
                st["left"]    # Observação
            ]

            # --- 3. LINHA 3 EM DIANTE: DADOS ---
            for i, mov in enumerate(movs, 3):

                vencido = False if mov.lote.data_vencimento is None else mov.lote.data_vencimento < hoje
                fill = st["vf"] if vencido else None
                font = st["vft"] if vencido else None
                
                _aplicar_linha(ws, i, [
                    mov.data_hora.strftime("%d/%m/%Y %H:%M"),
                    mov.lote.produto.nome,
                    mov.lote.num_lote,
                    mov.numero_nf or mov.lote.nota_fiscal,
                    mov.tipo.value.replace("_", " ").title(),
                    mov.quantidade,
                    mov.usuario.nome,
                    mov.observacao or "",
                ], st, fill=fill, font=font, alignments=alinhamentos_da_linha)
                
        # Congela as duas linhas superiores
        ws.freeze_panes = "A3"

        _buf = _io_bg.BytesIO()
        wb.save(_buf)
        _bytes_final = _aplicar_background(_buf.getvalue())
        caminho = XlsxBuilder._nome_arquivo("movimentacao")
        Path(caminho).write_bytes(_bytes_final)
        logger.info("Relatório movimentação gerado: %s", caminho)
        return caminho
 
    # ── 2. Estoque atual ───────────────────────────────────────────────────
 
    @staticmethod
    def estoque_atual() -> Path:
        """
        Posição atual do estoque por lote.
        RF-16: produto, lote, NF, datas, saldo, valor e situação.
        Lotes vencidos em vermelho (RF-22).
        """
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Estoque Atual"
        st = _wb_styles()
 
        colunas = [
            ("Produto", 35), ("Centro", 14), ("Lote", 14),
            ("Nota Fiscal", 14), ("Fabricação", 14), ("Vencimento", 14),
            ("Qtd. Inicial", 12), ("Qtd. Atual", 12),
            ("Vlr. Unit. (R$)", 14), ("Vlr. Total (R$)", 14), ("Situação", 16),
        ]

 
        hoje = date.today()
        with get_read_session() as s:
            lotes = (
                s.query(Lote)
                .join(Produto)
                .options(joinedload(Lote.produto))
                .filter(Produto.ativo == True, Lote.quantidade_atual > 0)
                .order_by(Produto.nome, Lote.data_vencimento)
                .all()
            )
             # --- 1. CONFIGURAÇÃO DA LINHA 1 (PERÍODO E TOTAL) ---
            ws.row_dimensions[1].height = 25
            
            # Mesclagem A1:D1 para o Período
            ws.merge_cells('A1:K1')
            cell_p = ws['A1']
            cell_p.value = f"Estoque {date.today()}"
            cell_p.font = Font(bold=True, color=COR_HEADER_FILL, size=11)
            cell_p.alignment = Alignment(horizontal="center", vertical="center")

            # --- 2. LINHA 2: TÍTULOS DAS COLUNAS ---
            _aplicar_header(ws, colunas, st, row_idx=2)
            for i, l in enumerate(lotes, 3):

                diff  = None if l.data_vencimento is None else (l.data_vencimento - hoje).days
                vencido = False if l.data_vencimento is None else l.data_vencimento < hoje
                if diff is None:
                    situacao = "Normal"
                    fill, font = None, None
                else:
                    if vencido:
                        situacao = "VENCIDO"
                        fill, font = st["vf"], st["vft"]
                    elif diff <= 7:
                        situacao = f"Vence em {diff}d"
                        fill, font = st["af"], st["aft"]
                    elif diff <= 15:
                        situacao = f"Vence em {diff}d"
                        fill, font = st["af"], st["aft"]
                    else:
                        situacao = "Normal"
                        fill, font = None, None
 
                _aplicar_linha(ws, i, [
                    l.produto.nome,
                    l.centro_alocacao.value.capitalize(),
                    l.num_lote,
                    l.nota_fiscal,
                    l.data_fabricacao.strftime("%d/%m/%Y") if l.data_fabricacao else "—",
                    l.data_vencimento.strftime("%d/%m/%Y") if l.data_vencimento else "—",
                    l.quantidade_inicial,
                    l.quantidade_atual,
                    float(l.valor_unitario),
                    float(l.valor_total),
                    situacao,
                ], st, fill=fill, font=font)
 
        ws.freeze_panes = "A3"
        _buf = _io_bg.BytesIO()
        wb.save(_buf)
        _bytes_final = _aplicar_background(_buf.getvalue())
        caminho = XlsxBuilder._nome_arquivo("estoque_atual")
        Path(caminho).write_bytes(_bytes_final)
        logger.info("Relatório estoque atual gerado: %s", caminho)
        return caminho
 
    # ── 3. Produtos a vencer em 30 dias ────────────────────────────────────
 
    @staticmethod
    def a_vencer(dias: int = 30) -> Path:
        """
        Lista lotes com vencimento nos próximos N dias (padrão 30).
        RF-17: ordenado por data de vencimento crescente.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"A Vencer ({dias}d)"
        st = _wb_styles()
 
        colunas = [
            ("Produto", 35), ("Centro", 14), ("Lote", 14),
            ("Nota Fiscal", 14), ("Vencimento", 14),
            ("Dias restantes", 14), ("Qtd. Atual", 12), ("Situação", 16),
        ]
 
        hoje   = date.today()
        limite = hoje + timedelta(days=dias)
 
        with get_read_session() as s:
            lotes = (
                s.query(Lote)
                .join(Produto)
                .options(joinedload(Lote.produto))
                .filter(
                    Produto.ativo == True,
                    Lote.quantidade_atual > 0,
                    Lote.data_vencimento >= hoje,
                    Lote.data_vencimento <= limite,
                    Lote.data_vencimento.isnot(None),
                )
                .order_by(Lote.data_vencimento)
                .all()
            )
            
            # Mesclagem A1:D1 para o Período
            ws.merge_cells('A1:H1')
            cell_p = ws['A1']
            cell_p.value = f"Lotes a vencer Proximos 30 Dias consulta {date.today()}"
            cell_p.font = Font(bold=True, color=COR_HEADER_FILL, size=11)
            cell_p.alignment = Alignment(horizontal="center", vertical="center")
            
            _aplicar_header(ws, colunas, st, row_idx=2)

            for i, l in enumerate(lotes, 3):
                diff = (l.data_vencimento - hoje).days
                if diff <= 2:
                    fill, font, sit = st["vf"], st["vft"], "Crítico"
                elif diff <= 7:
                    fill, font, sit = st["af"], st["aft"], "Urgente"
                else:
                    fill, font, sit = None, None, "Atenção"
 
                _aplicar_linha(ws, i, [
                    l.produto.nome,
                    l.centro_alocacao.value.capitalize(),
                    l.num_lote,
                    l.nota_fiscal,
                    l.data_vencimento.strftime("%d/%m/%Y"),
                    diff,
                    l.quantidade_atual,
                    sit,
                ], st, fill=fill, font=font)
 
        ws.freeze_panes = "A3"
        _buf = _io_bg.BytesIO()
        wb.save(_buf)
        _bytes_final = _aplicar_background(_buf.getvalue())
        caminho = XlsxBuilder._nome_arquivo("a_vencer")
        Path(caminho).write_bytes(_bytes_final)
        logger.info("Relatório a vencer gerado: %s (%d lotes)", caminho, len(lotes))
        return caminho
 
    # ── 4. Lotes vencidos em estoque ───────────────────────────────────────
 
    @staticmethod
    def lotes_vencidos() -> Path:
        """
        Lista lotes com data de vencimento anterior a hoje com saldo > 0.
        RF-22: todos destacados em vermelho.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Lotes Vencidos"
        st = _wb_styles()
 
        hoje = date.today()
        
        with get_read_session() as s:
            # 1. Busca os lotes primeiro para ter os dados de contagem e valor
            lotes = (
                s.query(Lote)
                .join(Produto)
                .options(joinedload(Lote.produto))
                .filter(
                    Produto.ativo == True,
                    Lote.quantidade_atual > 0,
                    Lote.data_vencimento < hoje,
                    Lote.data_vencimento.isnot(None),
                )
                .order_by(Lote.data_vencimento)
                .all()
            )
            if not lotes:
                logger.info("Nenhum lote vencido em estoque encontrado.")
                return None

            # 2. Configura a Top Bar (Linha 1) - Título com a data do dia
            ws.row_dimensions[1].height = 30
            ws.merge_cells('A1:L1')
            cell_titulo = ws['A1']
            cell_titulo.value = f"Lotes vencidos em estoque no dia {hoje.strftime('%d/%m/%Y')}"
            cell_titulo.font = Font(bold=True, color=COR_HEADER_FONT, size=12)
            cell_titulo.fill = PatternFill("solid", fgColor=COR_HEADER_FILL)
            cell_titulo.alignment = Alignment(horizontal="center", vertical="center")

            # 3. Cabeçalhos das Colunas (Linha 2, Colunas A até I)
            colunas = [
                ("Produto", 35), ("Centro", 14), ("Fornecedor", 25),
                ("Lote", 14), ("Nota Fiscal", 14), ("Vencimento", 14),
                ("Dias vencido", 14), ("Qtd. em estoque", 15),
                ("Valor em estoque (R$)", 22),
            ]
            _aplicar_header(ws, colunas, st, row_idx=2)

            valor_total_geral = sum(l.valor_unitario * l.quantidade_atual for l in lotes)
            
            cell_resumo = ws['J2']
            cell_resumo.value = (
                f"Total de Lotes: {len(lotes)} vencidos")
            
            # Estilo do quadro de resumo
            cell_resumo.font = Font(bold=True, color=COR_HEADER_FONT, size=11)
            cell_resumo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_resumo.border = st["border"]
            cell_resumo.fill = PatternFill("solid", fgColor=COR_HEADER_FILL)

            ws.merge_cells('J3:J4')
            cell_total= ws['J3']
            cell_total.value=(f"Valor: {valor_total_geral}")

            cell_total.font = Font(bold=True, color=COR_HEADER_FONT, size=11)
            cell_total.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_total.border = st["border"]
            cell_total.fill = PatternFill("solid", fgColor=COR_VENCIDO_FONT)

            ws.column_dimensions['J'].width = 30

            # 5. Preenchimento dos Dados (A partir da Linha 3)
            alinhamentos = [st["left"], st["center"], st["left"], st["center"], 
                            st["center"], st["center"], st["center"], st["center"], st["center"]]

            for i, l in enumerate(lotes, 3):
                dias_vencido = (hoje - l.data_vencimento).days
                valor_em_estoque = l.valor_unitario * l.quantidade_atual
 
                # Todos vencidos → destaque em vermelho (RF-22)
                _aplicar_linha(ws, i, [
                    l.produto.nome,
                    l.centro_alocacao.value.capitalize(),
                    l.produto.fornecedor or "—",
                    l.num_lote,
                    l.nota_fiscal,
                    l.data_vencimento.strftime("%d/%m/%Y"),
                    dias_vencido,
                    l.quantidade_atual,
                    float(valor_em_estoque),
                ], st, fill=st["vf"], font=st["vft"], alignments=alinhamentos)
 
        ws.freeze_panes = "A3" # Congela Título e Cabeçalho
        _buf = _io_bg.BytesIO()
        wb.save(_buf)
        _bytes_final = _aplicar_background(_buf.getvalue())
        caminho = XlsxBuilder._nome_arquivo("lotes_vencidos")
        Path(caminho).write_bytes(_bytes_final)
        logger.info("Relatório lotes vencidos gerado: %s (%d lotes)", caminho, len(lotes))
        return caminho

    # ── 5. Consumo médio por produto ────────────────────────────────────────

    @staticmethod
    def consumo_medio(dados: list[list], meses_labels: list[str]) -> Path:
        """
        Consumo (saídas) mês a mês, por grupo de consumo ou produto individual,
        nos últimos meses de calendário indicados em `meses_labels` (ex.:
        ["Jun/2026", "Jul/2026", "Ago/2026"]). `dados` já vem pronto de
        RelatorioService.buscar_dados_consumo_medio — cada linha é
        [rótulo, mês1, ..., mêsN, total, média_mensal].
        """
        meses = len(meses_labels)
        wb = Workbook()
        ws = wb.active
        ws.title = f"Consumo Médio ({meses}m)"
        st = _wb_styles()

        colunas = (
            [("Produto/Grupo", 100)]
            + [(label, 12) for label in meses_labels]
            + [("Total Consumido", 16), ("Média Mensal", 14)]
        )
        n_cols = len(colunas)
        ultima_letra = chr(ord('A') + n_cols - 1)

        ws.row_dimensions[1].height = 25
        ws.merge_cells(f'A1:{ultima_letra}1')
        cell_p = ws['A1']
        cell_p.value = (f"Consumo médio — últimos {meses} meses "
                         f"({meses_labels[0]} a {meses_labels[-1]}, base: saídas de estoque)")
        cell_p.font = Font(bold=True, color=COR_HEADER_FILL, size=11)
        cell_p.alignment = Alignment(horizontal="center", vertical="center")

        _aplicar_header(ws, colunas, st, row_idx=2)

        alinhamentos = [st["left"]] + [st["center"]] * (n_cols - 1)
        for i, linha in enumerate(dados, 3):
            _aplicar_linha(ws, i, linha, st, alignments=alinhamentos)

        # "A3" em vez de "B3": congela só as linhas 1-2 (título + cabeçalho),
        # sem congelar coluna nenhuma. Um freeze de coluna (ex. "B3") cruzaria
        # a célula mesclada do título (que ocupa A1 até a última coluna),
        # cortando-a visualmente ao rolar a planilha para os lados.
        ws.freeze_panes = "A3"
        _buf = _io_bg.BytesIO()
        wb.save(_buf)
        _bytes_final = _aplicar_background(_buf.getvalue())
        caminho = XlsxBuilder._nome_arquivo("consumo_medio")
        Path(caminho).write_bytes(_bytes_final)
        logger.info("Relatório consumo médio gerado: %s (%d linhas)", caminho, len(dados))
        return caminho

    # ── 6. Sessão de inventário (MOD-07) ────────────────────────────────────

    @staticmethod
    def relatorio_inventario(inventario, resumo, itens: list, sobras: list) -> Path:
        """
        Relatório de uma sessão de inventário (RF-34). Abas: resumo,
        encontrados, divergências, não localizados, sobras — os itens já
        vêm classificados por InventarioService, este método só distribui
        pelas abas.
        """
        wb = Workbook()
        st = _wb_styles()

        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        ws_resumo.merge_cells('A1:B1')
        cell_p = ws_resumo['A1']
        cell_p.value = f"Inventário #{inventario.id} — {inventario.descricao}"
        cell_p.font = Font(bold=True, color=COR_HEADER_FILL, size=12)
        cell_p.alignment = Alignment(horizontal="center", vertical="center")

        escopo_label = (
            "Geral" if inventario.escopo.value == "geral"
            else (inventario.localizacao.nome_completo if inventario.localizacao else "—")
        )
        linhas_resumo = [
            ("Escopo", escopo_label),
            ("Status", resumo.status.replace("_", " ").title()),
            ("Aberto em", inventario.aberto_em.strftime("%d/%m/%Y %H:%M")),
            ("Total esperado", resumo.total_esperado),
            ("Encontrados", resumo.encontrados),
            ("Divergentes", resumo.divergentes),
            ("Não localizados", resumo.nao_localizados),
            ("Sobras", resumo.sobras),
        ]
        _aplicar_header(ws_resumo, [("Indicador", 30), ("Valor", 40)], st, row_idx=2)
        for i, (rotulo, valor) in enumerate(linhas_resumo, 3):
            _aplicar_linha(ws_resumo, i, [rotulo, valor], st)
        ws_resumo.freeze_panes = "A3"

        def _sheet_itens(titulo: str, lista, fill=None, font=None):
            ws = wb.create_sheet(titulo)
            colunas = [
                ("Tombo", 16), ("Descrição", 45), ("Localização esperada", 28),
                ("Localização encontrada", 28), ("Conferido em", 18),
                ("Conferido por", 22), ("Observação", 40),
            ]
            _aplicar_header(ws, colunas, st, row_idx=1)
            for i, item in enumerate(lista, 2):
                _aplicar_linha(ws, i, [
                    item.bem.tombo,
                    item.bem.descricao,
                    item.localizacao_esperada.nome_completo if item.localizacao_esperada else "—",
                    item.localizacao_encontrada.nome_completo if item.localizacao_encontrada else "—",
                    item.conferido_em.strftime("%d/%m/%Y %H:%M") if item.conferido_em else "—",
                    item.conferido_por_usuario.nome if item.conferido_por_usuario else "—",
                    item.observacao or "",
                ], st, fill=fill, font=font)
            ws.freeze_panes = "A2"
            return ws

        encontrados      = [i for i in itens if i.status.value == "encontrado"]
        divergentes      = [i for i in itens if i.status.value == "divergente_local"]
        nao_localizados  = [i for i in itens if i.status.value == "nao_localizado"]

        _sheet_itens("Encontrados", encontrados)
        _sheet_itens("Divergências", divergentes, fill=st["af"], font=st["aft"])
        _sheet_itens("Não localizados", nao_localizados, fill=st["vf"], font=st["vft"])

        ws_sobras = wb.create_sheet("Sobras")
        colunas_sobras = [
            ("Código lido", 24), ("Tipo", 20), ("Localização", 28),
            ("Registrado em", 18), ("Registrado por", 22), ("Descrição livre", 40),
        ]
        _aplicar_header(ws_sobras, colunas_sobras, st, row_idx=1)
        for i, sobra in enumerate(sobras, 2):
            _aplicar_linha(ws_sobras, i, [
                sobra.codigo_lido,
                sobra.tipo.value.replace("_", " ").title(),
                sobra.localizacao.nome_completo if sobra.localizacao else "—",
                sobra.registrado_em.strftime("%d/%m/%Y %H:%M"),
                sobra.registrado_por_usuario.nome if sobra.registrado_por_usuario else "—",
                sobra.descricao_livre or "",
            ], st, fill=st["af"], font=st["aft"])
        ws_sobras.freeze_panes = "A2"

        _buf = _io_bg.BytesIO()
        wb.save(_buf)
        _bytes_final = _aplicar_background(_buf.getvalue())
        caminho = XlsxBuilder._nome_arquivo("inventario")
        Path(caminho).write_bytes(_bytes_final)
        logger.info("Relatório de inventário gerado: %s (inventario_id=%s)", caminho, inventario.id)
        return caminho